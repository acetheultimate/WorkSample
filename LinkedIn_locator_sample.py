from bs4 import BeautifulSoup
from selenium import webdriver
from openpyxl import load_workbook
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import NoSuchElementException
import random
import time


def generate_user_agent_and_proxy():
    pl = ['188.32.106.120:8081',
          '95.79.41.94:8081',
          '188.255.29.89:8081',
          '195.123.209.104:80',
          '36.67.50.242:8080',
          '36.228.41.168:8888',
          '175.139.65.229:8080',
          '187.87.77.76:3128',
          '223.19.210.69:80',
          '91.205.52.234:8081']
    user_agents = [
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/42.0.2311.135 Safari/537.36 Edge/12.246",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_2) AppleWebKit/601.3.9 (KHTML, like Gecko) Version/9.0.2 Safari/601.3.9",
        # "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/47.0.2526.111 Safari/537.36",
        "Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:15.0) Gecko/20100101 Firefox/15.0.1"]

    return random.choice(pl), random.choice(user_agents)

proxy_add, user_agent = generate_user_agent_and_proxy()
phantomjs_path = "./chromedriver"
service_args = [
    '--proxy=%s' % proxy_add,
    '--proxy-type=http',
    '--ssl-protocol=any',
    '--ignore-ssl-errors=true'
]
opts = webdriver.ChromeOptions()
opts.add_argument("--user-agent=%s" % user_agent)
driver = webdriver.Chrome(executable_path=phantomjs_path,
                          service_args=service_args,
                          chrome_options=opts
                          )

driver.set_window_size(1366, 768)
driver.get("https://www.linkedin.com/mynetwork/invite-connect/connections/")
frame = driver.find_element_by_class_name("authentication-iframe")
driver.switch_to.frame(frame)
driver.find_element_by_class_name("sign-in-link").click()

# Logging in with LinkedIn creds
driver.find_element_by_id("session_key-login").send_keys("yourid@domain.com")
driver.find_element_by_id("session_password-login").send_keys("your_password")
driver.find_element_by_id("btn-primary").click()

# Wait for the session to be stored on LinkedIn
time.sleep(5)


def scrapper(linkedin, nr, driver):
    linkedin = linkedin.strip()
    print("Opening ", linkedin)

    # A simple javascript hack to preserve the session ;)
    # Instead of setting URL on driver, which probably can delete the session, Just change the href.location
    driver.execute_script('window.location.href = "%s"' % linkedin)
    print("Opened!")

    output_file = load_workbook("output.xlsx")
    output_sheet = output_file.active

    print("souping it up!")
    soup = BeautifulSoup(driver.page_source, 'html.parser')

    # Declaration of variables needed
    name = None
    location = None
    company = None
    desig = None
    email = None
    phone = None

    # agent = driver.execute_script("return navigator.userAgent")
    # print("Agent was", agent)
    try:
        element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, "pv-top-card-section__body"))
        )
        name = element.find_element_by_class_name("pv-top-card-section__name").text.strip()
        location = element.find_element_by_class_name("pv-top-card-section__location").text.strip()
        try:
            company = element.find_element_by_class_name("pv-top-card-section__company").text.strip()
        except NoSuchElementException:
            company = ""

        if company == "":
            try:
                company = [i.strip() for i in element.find_element_by_class_name("pv-top-card-section__headline").text.split(" at ")][1]
            except IndexError:
                company = ""
        try:
            desig = [i.strip() for i in element.find_element_by_class_name("pv-top-card-section__headline").text.split(" at ")][0]
        except ValueError:
            desig = element.find_element_by_class_name("pv-top-card-section__headline").text.strip()

        if ("student" in desig.lower()) or ("intern" in desig.lower()):
            return True

    except TimeoutException:
        driver.quit()

    try:
        button = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, "contact-see-more-less"))
        )
        button.click()
        element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, "ci-email"))
        )
        try:
            email = element.find_element_by_class_name("pv-contact-info__contact-item").text.strip()
        except NoSuchElementException:
            email = ''
        try:
            phone = driver.find_element_by_class_name("ci-phone").find_element_by_class_name("pv-contact-info__contact-item").text.strip()
            phone = phone.split("(")[0].strip()
        except NoSuchElementException:
            phone = ''
    except TimeoutException:
        pass
    nr += 1

    # print(name, email, phone, company, desig, location)
    output_sheet.cell(row=nr, column=1).value = name
    output_sheet.cell(row=nr, column=2).value = email
    output_sheet.cell(row=nr, column=3).value = phone
    output_sheet.cell(row=nr, column=4).value = company
    output_sheet.cell(row=nr, column=5).value = desig
    output_sheet.cell(row=nr, column=6).value = location
    output_sheet.cell(row=nr, column=7).value = linkedin
    output_file.save('output.xlsx')
    return True


def begin(driver):
    # Read lines containing links to the linkedin users
    for url in enumerate(open("raw_out.txt").readlines()):
        if scrapper(url[1], url[0], driver):
            print("Request Completed for", url[1], "at", url[0])
        else:
            print("Some error at", url[0])
            break
begin(driver)

driver.quit()
